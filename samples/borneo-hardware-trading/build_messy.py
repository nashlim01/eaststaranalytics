"""Build the 'BEFORE' file: messy_sales_2025.xlsx — what the client currently lives with.
12 monthly tabs, inconsistent everything. Derived from data.json but deliberately degraded."""
import json, random, datetime as dt
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

random.seed(7)
HERE = Path(__file__).parent
data = json.loads((HERE/"data.json").read_text())
records = data["records"]
MONTHS = data["meta"]["months"]

# Inconsistent tab names — the kind of thing that makes consolidation a nightmare
TAB_NAMES = ["Jan","FEB 2025","March","Apr-25","MAY","June 2025","Jul","August",
             "Sept","Oct 2025","NOV","Dec"]
MALAY_MON = {1:"Jan",2:"Feb",3:"Mac",4:"Apr",5:"Mei",6:"Jun",7:"Jul",8:"Ogos",9:"Sep",10:"Okt",11:"Nov",12:"Dis"}

# Rep name corruption — same person, many spellings
REP_VARIANTS = {
    "Ahmad Faizal":  ["Ahmad Faizal","ahmad faizal","A. Faizal","Ahmad","AHMAD FAIZAL"],
    "Lim Ah Hock":   ["Lim Ah Hock","Ah Hock","ah hock","Lim AH","A.Hock"],
    "Rosli bin Osman":["Rosli bin Osman","Rosli","rosli osman","R. Osman"],
    "Catherine Wong":["Catherine Wong","Cathy","catherine wong","C. Wong","Catherine"],
    "Mohd Idris":    ["Mohd Idris","Idris","mohd idris","M.Idris","MOHD IDRIS"],
}
PROD_VARIANTS = {
    "OPC Cement 50kg": ["OPC Cement 50kg","Cement 50kg","cement","OPC cement","Cement(50kg)"],
    "Steel Bar Y12":   ["Steel Bar Y12","Y12 bar","steel y12","Bar Y12"],
    "Steel Bar Y16":   ["Steel Bar Y16","Y16","steel bar y16","Bar Y16 "],
    "Plywood 18mm":    ["Plywood 18mm","plywood 18","Ply 18mm","18mm plywood"],
    "Marine Plywood 12mm":["Marine Plywood 12mm","marine ply 12","Marine PW 12mm"],
    "Roofing Sheet 0.4mm":["Roofing Sheet 0.4mm","roofing 0.4","Roof sheet","Zinc roofing 0.4"],
    "PVC Pipe 4in":    ["PVC Pipe 4in","pvc 4\"","PVC pipe 4inch","4in PVC"],
    "Emulsion Paint 18L":["Emulsion Paint 18L","paint 18L","Emulsion 18l","Paint(18L)"],
    "Cordless Drill":  ["Cordless Drill","drill","Cordless drill ","Power Drill"],
    "Wire Nails 50kg": ["Wire Nails 50kg","nails 50kg","wire nail","Nails"],
    "Sand (per ton)":  ["Sand (per ton)","sand","Sand/ton","River Sand"],
    "Red Brick (per 1000)":["Red Brick (per 1000)","brick 1000","Red brick","Bricks"],
}

def messy_date(iso, mi):
    d = dt.date.fromisoformat(iso)
    style = random.random()
    if style < 0.45:        # real date object
        return d, "DD/MM/YYYY"
    elif style < 0.70:      # text dd/mm/yyyy
        return f"{d.day}/{d.month}/{d.year}", None
    elif style < 0.88:      # malay text date
        return f"{d.day} {MALAY_MON[d.month]} {d.year}", None
    else:                   # text with dashes
        return f"{d.day:02d}-{d.month:02d}-{d.year}", None

def messy_amount(amt):
    style = random.random()
    if style < 0.55:
        return amt, "#,##0.00"          # plain number
    elif style < 0.80:
        return f"RM {amt:,.2f}", None    # text with RM
    elif style < 0.92:
        return f"RM{amt:.2f}", None      # text no space
    else:
        return amt, "0.00"               # number, no thousands sep

wb = Workbook()
wb.remove(wb.active)

title_fill = PatternFill("solid", fgColor="C9D9C0")
hdr_fill = PatternFill("solid", fgColor="E8E8E8")
thin = Side(style="thin", color="BBBBBB")
border = Border(left=thin,right=thin,top=thin,bottom=thin)
ARIAL = "Arial"

for mi, mon in enumerate(MONTHS):
    rows = [r for r in records if r["month"] == mon]
    ws = wb.create_sheet(TAB_NAMES[mi][:31])
    # Merged title banner (inconsistent wording)
    titles = [f"Borneo Hardware - Sales {mon} 2025", f"SALES RECORD {mon.upper()}",
              f"{mon} 2025 sales", f"Jualan {MALAY_MON[mi+1]} 2025"]
    ws.merge_cells("A1:F1")
    t = ws["A1"]; t.value = random.choice(titles)
    t.font = Font(name=ARIAL, bold=True, size=13); t.fill = title_fill
    t.alignment = Alignment(horizontal="center")

    # Header row — column order/labels drift slightly between sheets
    header_variants = [
        ["Date","Customer","Item","Qty","Amount","Sales Person"],
        ["Date","Customer","Product","Qty","Amount (RM)","Salesperson"],
        ["Tarikh","Customer","Item","Unit","Amount","Sales"],
    ]
    hdr = random.choice(header_variants)
    hrow = 3
    for c, h in enumerate(hdr, start=1):
        cell = ws.cell(row=hrow, column=c, value=h)
        cell.font = Font(name=ARIAL, bold=True); cell.fill = hdr_fill; cell.border = border

    r = hrow + 1
    running_total = 0.0
    for rec in rows:
        # occasional fully blank separator row
        if random.random() < 0.04:
            r += 1
        dval, dfmt = messy_date(rec["date"], mi)
        aval, afmt = messy_amount(rec["amount"])
        running_total += rec["amount"]
        rep = random.choice(REP_VARIANTS[rec["rep"]])
        prod = random.choice(PROD_VARIANTS[rec["product"]])
        cust = rec["customer"]
        if random.random() < 0.05: cust = ""        # missing customer
        if random.random() < 0.03: prod = prod.upper()
        ws.cell(row=r, column=1, value=dval)
        if dfmt: ws.cell(row=r, column=1).number_format = dfmt
        ws.cell(row=r, column=2, value=cust)
        ws.cell(row=r, column=3, value=prod)
        ws.cell(row=r, column=4, value=rec["qty"])
        ac = ws.cell(row=r, column=5, value=aval)
        if afmt: ac.number_format = afmt
        ws.cell(row=r, column=6, value=rep)
        for c in range(1,7):
            ws.cell(row=r, column=c).font = Font(name=ARIAL, size=10)
        r += 1

    # Hand-typed TOTAL row — sometimes hardcoded & slightly wrong (transcription error)
    r += 1
    ws.cell(row=r, column=3, value="TOTAL").font = Font(name=ARIAL, bold=True)
    typed = running_total
    if random.random() < 0.35:   # someone fat-fingered the manual sum
        typed = round(running_total * random.uniform(0.97, 1.03), 2)
    tc = ws.cell(row=r, column=5, value=round(typed,2))
    tc.font = Font(name=ARIAL, bold=True); tc.number_format = "#,##0.00"

    widths = [16,26,20,7,14,16]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.sheet_view.showGridLines = True

out = HERE/"1_messy_sales_2025_BEFORE.xlsx"
wb.save(out)
print("saved", out.name, "| 12 tabs:", ", ".join(TAB_NAMES))
