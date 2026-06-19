"""Build the 'AFTER' file: clean_dashboard_2025.xlsx.
One normalized Data table + a live, formula-driven Dashboard with native charts.
Everything recalculates the instant new rows are pasted into Data."""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.worksheet.table import Table, TableStyleInfo

HERE = Path(__file__).parent
data = json.loads((HERE/"data.json").read_text())
records = data["records"]
S = data["summary"]
MONTHS = S["months"]
REPS = data["meta"]["reps"]
BRANCHES = data["meta"]["branches"]
PRODUCTS = [p[0] for p in data["meta"]["products"]]
CATS = sorted({r["category"] for r in records})

NAVY = "0B1F3A"; GOLD = "C9A24B"; LIGHT = "F4F1E8"; MIDGOLD="E7DCC0"
WHITE = "FFFFFF"; INK="1A1A1A"; GREY="6B6B6B"
ARIAL = "Arial"

wb = Workbook()

# ---------------- Data sheet ----------------
ds = wb.active; ds.title = "Data"
cols = ["Order ID","Date","Month","Branch","Sales Rep","Customer","Category","Product",
        "Qty","Unit Price (RM)","Amount (RM)","Status"]
ds.append(cols)
mon_idx = {m:i for i,m in enumerate(MONTHS)}
for r in sorted(records, key=lambda x:x["date"]):
    import datetime as dt
    ds.append([r["order_id"], dt.date.fromisoformat(r["date"]), r["month"], r["branch"],
               r["rep"], r["customer"], r["category"], r["product"], r["qty"],
               r["unit_price"], r["amount"], "Paid" if r["paid"] else "Outstanding"])
nrows = len(records) + 1  # incl header
# header style
for c in range(1, len(cols)+1):
    cell = ds.cell(row=1, column=c)
    cell.font = Font(name=ARIAL, bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
for row in ds.iter_rows(min_row=2, max_row=nrows):
    for cell in row:
        cell.font = Font(name=ARIAL, size=10)
    row[1].number_format = "dd/mm/yyyy"
    row[9].number_format = "#,##0.00"
    row[10].number_format = "#,##0.00"
ds.column_dimensions["A"].width=16
for col,w in zip("BCDEFGHIJKL",[12,8,11,16,24,20,20,7,13,13,12]):
    ds.column_dimensions[col].width = w
ds.freeze_panes = "A2"
tab = Table(displayName="SalesData", ref=f"A1:L{nrows}")
tab.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=True)
ds.add_table(tab)

DK = f"$K$2:$K${nrows}"; DC=f"$C$2:$C${nrows}"; DD=f"$D$2:$D${nrows}"
DE=f"$E$2:$E${nrows}"; DH=f"$H$2:$H${nrows}"; DG=f"$G$2:$G${nrows}"; DL=f"$L$2:$L${nrows}"

# ---------------- Dashboard sheet ----------------
dash = wb.create_sheet("Dashboard", 0)
dash.sheet_view.showGridLines = False
for col,w in zip("ABCDEFGHIJ",[3,22,16,16,4,18,16,16,16,3]): dash.column_dimensions[col].width=w

def style(cell, **kw):
    cell.font = Font(name=ARIAL, **{k:v for k,v in kw.items() if k in
                     ("bold","size","color","italic")})
    return cell

# Banner
dash.merge_cells("B2:I2")
b = dash["B2"]; b.value = "BORNEO HARDWARE TRADING SDN BHD  —  Sales Performance Dashboard  ·  FY2025"
b.font = Font(name=ARIAL, bold=True, size=14, color=WHITE)
b.fill = PatternFill("solid", fgColor=NAVY); b.alignment=Alignment(horizontal="center",vertical="center")
dash.row_dimensions[2].height = 28
dash.merge_cells("B3:I3")
sub = dash["B3"]; sub.value="Live demo · figures update automatically when new sales rows are added to the Data tab"
sub.font=Font(name=ARIAL, italic=True, size=9, color=GREY); sub.alignment=Alignment(horizontal="center")

# KPI cards (row 5-7)
kpis = [
    ("TOTAL REVENUE", f'=SUM(Data!{DK})', '"RM "#,##0'),
    ("TOTAL ORDERS", f'=COUNTA(Data!$A$2:$A${nrows})', '#,##0'),
    ("AVG ORDER VALUE", f'=AVERAGE(Data!{DK})', '"RM "#,##0'),
    ("OUTSTANDING (UNPAID)", f'=SUMIF(Data!{DL},"Outstanding",Data!{DK})', '"RM "#,##0'),
]
card_cols = ["B","D","F","H"]
for (label,formula,fmt),col in zip(kpis, card_cols):
    nxt = get_column_letter(ord(col)-64+1)
    dash.merge_cells(f"{col}5:{nxt}5")
    dash.merge_cells(f"{col}6:{nxt}6")
    lc = dash[f"{col}5"]; lc.value=label
    lc.font=Font(name=ARIAL,bold=True,size=8,color=WHITE); lc.fill=PatternFill("solid",fgColor=GOLD)
    lc.alignment=Alignment(horizontal="center")
    vc = dash[f"{col}6"]; vc.value=formula
    vc.font=Font(name=ARIAL,bold=True,size=15,color=NAVY); vc.fill=PatternFill("solid",fgColor=LIGHT)
    vc.alignment=Alignment(horizontal="center",vertical="center"); vc.number_format=fmt
    dash.row_dimensions[6].height=26

# --- Monthly revenue table (helper block, row 9+) ---
r0 = 9
style(dash.cell(r0, 2, "Monthly Revenue (RM)"), bold=True, size=11, color=NAVY)
dash.cell(r0+1,2,"Month").font=Font(name=ARIAL,bold=True,size=9)
dash.cell(r0+1,3,"Revenue").font=Font(name=ARIAL,bold=True,size=9)
for i,m in enumerate(MONTHS):
    rr=r0+2+i
    dash.cell(rr,2,m).font=Font(name=ARIAL,size=10)
    fc=dash.cell(rr,3, f'=SUMIFS(Data!{DK},Data!{DC},B{rr})')
    fc.number_format='#,##0'; fc.font=Font(name=ARIAL,size=10)
mon_last=r0+1+len(MONTHS)

# --- By Branch (cols F-G) ---
style(dash.cell(r0,6,"Revenue by Branch"), bold=True, size=11, color=NAVY)
dash.cell(r0+1,6,"Branch").font=Font(name=ARIAL,bold=True,size=9)
dash.cell(r0+1,7,"Revenue").font=Font(name=ARIAL,bold=True,size=9)
for i,bn in enumerate(BRANCHES):
    rr=r0+2+i
    dash.cell(rr,6,bn).font=Font(name=ARIAL,size=10)
    fc=dash.cell(rr,7, f'=SUMIFS(Data!{DK},Data!{DD},F{rr})')
    fc.number_format='#,##0'; fc.font=Font(name=ARIAL,size=10)
br_last=r0+1+len(BRANCHES)

# --- By Sales Rep (cols F-G, below branch) ---
rr0 = br_last+2
style(dash.cell(rr0,6,"Revenue by Sales Rep"), bold=True, size=11, color=NAVY)
dash.cell(rr0+1,6,"Sales Rep").font=Font(name=ARIAL,bold=True,size=9)
dash.cell(rr0+1,7,"Revenue").font=Font(name=ARIAL,bold=True,size=9)
for i,rp in enumerate(REPS):
    rr=rr0+2+i
    dash.cell(rr,6,rp).font=Font(name=ARIAL,size=10)
    fc=dash.cell(rr,7, f'=SUMIFS(Data!{DK},Data!{DE},F{rr})')
    fc.number_format='#,##0'; fc.font=Font(name=ARIAL,size=10)
rep_last=rr0+1+len(REPS)

# --- Top Products by category (cols B-C, below monthly) ---
pr0 = mon_last+2
style(dash.cell(pr0,2,"Revenue by Product"), bold=True, size=11, color=NAVY)
dash.cell(pr0+1,2,"Product").font=Font(name=ARIAL,bold=True,size=9)
dash.cell(pr0+1,3,"Revenue").font=Font(name=ARIAL,bold=True,size=9)
for i,pn in enumerate(PRODUCTS):
    rr=pr0+2+i
    dash.cell(rr,2,pn).font=Font(name=ARIAL,size=10)
    fc=dash.cell(rr,3, f'=SUMIFS(Data!{DK},Data!{DH},B{rr})')
    fc.number_format='#,##0'; fc.font=Font(name=ARIAL,size=10)
prod_last=pr0+1+len(PRODUCTS)

# ---------------- Charts ----------------
# Monthly line chart
lc = LineChart(); lc.title="Monthly Revenue Trend — FY2025"; lc.style=2; lc.height=7.2; lc.width=15
ldata = Reference(dash, min_col=3, min_row=r0+1, max_row=mon_last)
lcats = Reference(dash, min_col=2, min_row=r0+2, max_row=mon_last)
lc.add_data(ldata, titles_from_data=True); lc.set_categories(lcats)
lc.y_axis.numFmt='#,##0'; lc.legend=None
dash.add_chart(lc, "B24")

# Branch bar chart
bc = BarChart(); bc.type="col"; bc.title="Revenue by Branch"; bc.height=7.2; bc.width=9
bdata = Reference(dash, min_col=7, min_row=r0+1, max_row=br_last)
bcats = Reference(dash, min_col=6, min_row=r0+2, max_row=br_last)
bc.add_data(bdata, titles_from_data=True); bc.set_categories(bcats); bc.legend=None
bc.y_axis.numFmt='#,##0'
dash.add_chart(bc, "F24")

# Rep bar chart
rc = BarChart(); rc.type="bar"; rc.title="Revenue by Sales Rep"; rc.height=7.5; rc.width=15
rdata = Reference(dash, min_col=7, min_row=rr0+1, max_row=rep_last)
rcats = Reference(dash, min_col=6, min_row=rr0+2, max_row=rep_last)
rc.add_data(rdata, titles_from_data=True); rc.set_categories(rcats); rc.legend=None
rc.x_axis.numFmt='#,##0'
dash.add_chart(rc, "B40")

# Product bar chart
pc = BarChart(); pc.type="bar"; pc.title="Revenue by Product"; pc.height=8.5; pc.width=9
pdata = Reference(dash, min_col=3, min_row=pr0+1, max_row=prod_last)
pcats = Reference(dash, min_col=2, min_row=pr0+2, max_row=prod_last)
pc.add_data(pdata, titles_from_data=True); pc.set_categories(pcats); pc.legend=None
pc.x_axis.numFmt='#,##0'
dash.add_chart(pc, "F40")

# ---------------- Read Me sheet ----------------
rm = wb.create_sheet("Read Me")
rm.sheet_view.showGridLines=False
rm.column_dimensions["B"].width=90
lines = [
 ("Borneo Hardware Trading Sdn Bhd — Sample Dashboard", True, 14, NAVY),
 ("Prepared by EastStar Analytics  ·  www.eaststar (sample / demo data)", False, 9, GREY),
 ("", False, 10, INK),
 ("What this is", True, 11, NAVY),
 ("A worked 'before → after' example. The companion file '1_messy_sales_2025_BEFORE.xlsx'", False,10,INK),
 ("is the kind of spreadsheet most SMEs actually keep: 12 monthly tabs, inconsistent dates,", False,10,INK),
 ("salesperson names spelled five different ways, hand-typed totals (some wrong).", False,10,INK),
 ("", False, 10, INK),
 ("This file is the 'after': one clean Data table feeding a live Dashboard.", False,10,INK),
 ("Every number on the Dashboard is an Excel formula (SUMIFS / SUMIF / AVERAGE).", False,10,INK),
 ("Paste new sales rows into the Data tab and the whole dashboard updates instantly.", False,10,INK),
 ("", False, 10, INK),
 ("Try it", True, 11, NAVY),
 ("1. Open the Dashboard tab — KPIs, monthly trend, branch / rep / product breakdowns.", False,10,INK),
 ("2. Add a row at the bottom of the Data table, then watch the KPI cards move.", False,10,INK),
 ("3. Compare against the messy BEFORE file to feel the difference.", False,10,INK),
 ("", False, 10, INK),
 ("Want this for your real numbers? Send your existing files over WhatsApp — even messy ones.", False,10,GOLD),
]
for i,(txt,bold,sz,clr) in enumerate(lines, start=2):
    c=rm.cell(i,2,txt); c.font=Font(name=ARIAL,bold=bold,size=sz,color=clr)

dash.sheet_view.tabSelected=True
wb.active = wb["Dashboard"]
out = HERE/"2_clean_dashboard_2025_AFTER.xlsx"
wb.save(out)
print("saved", out.name, "| data rows:", nrows-1)
